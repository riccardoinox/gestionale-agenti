import os
import openpyxl
import datetime
from typing import Dict, Any, Tuple
from database import get_db_connection, init_db

def safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        if isinstance(val, (int, float)):
            return float(val)
        cleaned = str(val).replace("€", "").replace(" ", "").replace(",", ".")
        return float(cleaned)
    except Exception:
        return default

def safe_int(val, default=0) -> int:
    if val is None:
        return default
    try:
        return int(float(val))
    except Exception:
        return default

def safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d/%m/%Y")
    return str(val).strip()

def safe_date_iso(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try parsing common Italian date formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s

EXCLUDED_ARTICLE_PREFIXES = (
    "BANC",
    "ESTINTORI",
    "L430",
    "LEGGE",
    "MATD",
    "PROVA",
    "SCONTO",
    "SPE",
)

def is_excluded_article(code: str) -> bool:
    if not code:
        return True
    code_upper = code.strip().upper()
    return any(code_upper.startswith(prefix) for prefix in EXCLUDED_ARTICLE_PREFIXES)

def is_excluded_client(name: str, code: str = "") -> bool:
    name_clean = str(name or "").strip()
    code_clean = str(code or "").strip()
    return name_clean.startswith("$") or code_clean.startswith("$")

def import_all_excel_data(base_dir: str = None) -> Dict[str, Any]:
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    init_db()
    conn = get_db_connection()
    cursor = conn.cursor()

    anagra_path = os.path.join(base_dir, "ANAGRA.xlsx")
    artico_path = os.path.join(base_dir, "ARTICO.xlsx")
    seor_path = os.path.join(base_dir, "SEOR.xlsx")
    
    # Check for Nuovo Listino file (which may have varying versions in filename)
    listino_path = None
    for fname in os.listdir(base_dir):
        if fname.lower().startswith("nuovo listino") and fname.endswith(".xlsx"):
            listino_path = os.path.join(base_dir, fname)
            break
    if not listino_path:
        listino_path = os.path.join(base_dir, "NUOVO LISTINO server ver.05.2026.xlsx")

    # 1. Read Prices from Listino
    price_map: Dict[str, float] = {}
    if os.path.exists(listino_path):
        print(f"Reading Listino from: {listino_path}")
        wb_list = openpyxl.load_workbook(listino_path, read_only=True, data_only=True)
        # Use BUSINESS sheet if exists, otherwise first sheet
        target_sheet = "BUSINESS" if "BUSINESS" in wb_list.sheetnames else wb_list.sheetnames[0]
        ws_list = wb_list[target_sheet]
        
        for row in ws_list.iter_rows(values_only=True):
            if not row or len(row) < 9:
                continue
            code = safe_str(row[1])
            if code and code.lower() != "descrizione":
                if not is_excluded_article(code):
                    price = safe_float(row[8], 0.0)
                    price_map[code.upper()] = price
        wb_list.close()
    print(f"Loaded {len(price_map)} prices from listino.")

    # 2. Read and Import Clients (Tab1.xlsx or ANAGRA.xlsx fallback)
    total_clients = 0
    tab1_path = os.path.join(base_dir, "Tab1.xlsx")
    
    if os.path.exists(tab1_path):
        print(f"Importing Clients from Tab1.xlsx: {tab1_path}")
        wb_tab1 = openpyxl.load_workbook(tab1_path, read_only=True, data_only=True)
        ws_tab1 = wb_tab1.active
        
        client_rows = []
        for i, row in enumerate(ws_tab1.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            name = safe_str(row[0])
            code = safe_str(row[8]) if len(row) > 8 else ""
            if not code or not name or code.lower() == "conto":
                continue
            
            # Exclude clients whose name or code starts with $
            if is_excluded_client(name, code):
                continue

            province = safe_str(row[1]).upper() if len(row) > 1 else ""
            address = safe_str(row[2]) if len(row) > 2 else ""
            city = safe_str(row[3]) if len(row) > 3 else ""
            email = safe_str(row[4]) if len(row) > 4 else ""
            phone = safe_str(row[5]) if len(row) > 5 else ""
            date_acq = safe_date_iso(row[6]) if len(row) > 6 else ""
            contact = safe_str(row[7]) if len(row) > 7 else ""
            agent_name = safe_str(row[9]) if len(row) > 9 else ""
            cap = safe_str(row[10]) if len(row) > 10 else ""
            mobile = safe_str(row[11]) if len(row) > 11 else ""

            client_rows.append((
                code, name, "", city, province, address, cap, email, agent_name,
                mobile, phone, "", "", "", contact, "", "", "", date_acq
            ))

        cursor.execute("DELETE FROM clients")
        cursor.executemany("""
            INSERT OR REPLACE INTO clients (
                code, name, name2, city, province, address, cap, email, agent_name,
                mobile, phone, fax, vat, tax_code, contact, first_name, last_name, subject_type, date_acq
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, client_rows)
        total_clients = len(client_rows)
        wb_tab1.close()
    elif os.path.exists(anagra_path):
        print(f"Importing Clients from legacy ANAGRA.xlsx: {anagra_path}")
        wb_anagra = openpyxl.load_workbook(anagra_path, read_only=True, data_only=True)
        ws_anagra = wb_anagra.active
        
        client_rows = []
        for i, row in enumerate(ws_anagra.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            code = safe_str(row[0])
            if not code or code.lower() == "codice":
                continue
            
            name = safe_str(row[1])
            if is_excluded_client(name, code):
                continue

            name2 = safe_str(row[2]) if len(row) > 2 else ""
            city = safe_str(row[3]) if len(row) > 3 else ""
            mobile = safe_str(row[4]) if len(row) > 4 else ""
            phone = safe_str(row[5]) if len(row) > 5 else ""
            fax = safe_str(row[6]) if len(row) > 6 else ""
            vat = safe_str(row[7]) if len(row) > 7 else ""
            tax_code = safe_str(row[8]) if len(row) > 8 else ""
            contact = safe_str(row[9]) if len(row) > 9 else ""
            subject_type = safe_str(row[14]) if len(row) > 14 else ""
            first_name = safe_str(row[15]) if len(row) > 15 else ""
            last_name = safe_str(row[16]) if len(row) > 16 else ""

            client_rows.append((
                code, name, name2, city, "", "", "", "", "",
                mobile, phone, fax, vat, tax_code, contact, first_name, last_name, subject_type, ""
            ))

        cursor.execute("DELETE FROM clients")
        cursor.executemany("""
            INSERT OR REPLACE INTO clients (
                code, name, name2, city, province, address, cap, email, agent_name,
                mobile, phone, fax, vat, tax_code, contact, first_name, last_name, subject_type, date_acq
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, client_rows)
        total_clients = len(client_rows)
        wb_anagra.close()
    print(f"Imported {total_clients} clients.")

    # 3. Read and Import Articles (ARTICO)
    total_articles = 0
    if os.path.exists(artico_path):
        print(f"Importing Articles from: {artico_path}")
        wb_artico = openpyxl.load_workbook(artico_path, read_only=True, data_only=True)
        ws_artico = wb_artico.active
        
        article_rows = []
        for i, row in enumerate(ws_artico.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            code = safe_str(row[0])
            if not code or code.lower() == "codice":
                continue

            if is_excluded_article(code):
                continue

            description = safe_str(row[1])
            disp_netta = safe_float(row[2], 0.0)
            ordinato = safe_float(row[3], 0.0)
            prenotato = safe_float(row[4], 0.0)
            impegnato = safe_float(row[5], 0.0)
            esistenza = safe_float(row[6], 0.0)
            esistenza_conv = safe_float(row[7], 0.0)
            es_imp = safe_float(row[8], 0.0)
            cod_altern = safe_str(row[9]) if len(row) > 9 else ""
            um = safe_str(row[10]) if len(row) > 10 else "PZ"
            disponib = safe_float(row[11], 0.0) if len(row) > 11 else 0.0
            ultimo_costo = safe_float(row[12], 0.0) if len(row) > 12 else 0.0
            conv = safe_float(row[13], 1.0) if len(row) > 13 else 1.0
            descr2 = safe_str(row[14]) if len(row) > 14 else ""
            art_sostitutivo = safe_str(row[17]) if len(row) > 17 else ""
            art_sostituito = safe_str(row[18]) if len(row) > 18 else ""
            in_esaurim = safe_str(row[19]).upper() if len(row) > 19 else "N"
            a_listino = safe_str(row[20]).upper() if len(row) > 20 else "S"

            price = price_map.get(code.upper(), 0.0)

            article_rows.append((
                code, description, disp_netta, ordinato, prenotato, impegnato,
                esistenza, esistenza_conv, es_imp, cod_altern, um, disponib,
                ultimo_costo, conv, descr2, art_sostitutivo, art_sostituito,
                in_esaurim, a_listino, price
            ))

        cursor.execute("DELETE FROM articles")
        cursor.executemany("""
            INSERT OR REPLACE INTO articles (
                code, description, disp_netta, ordinato, prenotato, impegnato,
                esistenza, esistenza_conv, es_imp, cod_altern, um, disponib,
                ultimo_costo, conv, descr2, art_sostitutivo, art_sostituito,
                in_esaurim, a_listino, listino_prezzo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, article_rows)
        total_articles = len(article_rows)
        wb_artico.close()
    print(f"Imported {total_articles} articles.")

    # 4. Read and Import Orders (SEOR)
    total_orders = 0
    if os.path.exists(seor_path):
        print(f"Importing Orders from: {seor_path}")
        wb_seor = openpyxl.load_workbook(seor_path, read_only=True, data_only=True)
        ws_seor = wb_seor.active
        
        order_rows = []
        for i, row in enumerate(ws_seor.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            year = safe_int(row[0], 2026)
            series = safe_str(row[1])
            number = safe_int(row[2], 0)
            if number == 0:
                continue

            order_id = f"{year}_{series}_{number}".replace(" ", "")
            order_date = safe_date_iso(row[3])
            client_code = safe_str(row[4])
            client_name = safe_str(row[5])
            delivery_date = safe_date_iso(row[6])
            total_amount = safe_float(row[7], 0.0)
            evaso = safe_str(row[8]).upper() if len(row) > 8 else "N"
            confermato = safe_str(row[9]).upper() if len(row) > 9 else "N"
            dest_code = safe_str(row[10]) if len(row) > 10 else ""
            dest_desc = safe_str(row[11]) if len(row) > 11 else ""
            reference = safe_str(row[14]) if len(row) > 14 else ""
            doc_type = safe_str(row[20]) if len(row) > 20 else ""
            aperto = safe_str(row[23]).upper() if len(row) > 23 else "N"
            sospeso = safe_str(row[24]).upper() if len(row) > 24 else "N"
            warehouse = safe_str(row[26]) if len(row) > 26 else ""

            order_rows.append((
                order_id, year, series, number, order_date, client_code, client_name,
                delivery_date, total_amount, evaso, confermato, dest_code, dest_desc,
                reference, doc_type, aperto, sospeso, warehouse
            ))

        cursor.execute("DELETE FROM orders")
        cursor.executemany("""
            INSERT OR REPLACE INTO orders (
                id, year, series, number, order_date, client_code, client_name,
                delivery_date, total_amount, evaso, confermato, dest_code, dest_desc,
                reference, doc_type, aperto, sospeso, warehouse
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, order_rows)
        total_orders = len(order_rows)
        wb_seor.close()
    print(f"Imported {total_orders} orders.")

    # 5. Read and Import Transports (TRASPORTI_2024.xlsx or similar)
    total_transports = 0
    trasporti_path = None
    for fname in os.listdir(base_dir):
        if fname.lower().startswith("trasporti") and fname.endswith(".xlsx"):
            trasporti_path = os.path.join(base_dir, fname)
            break

    if trasporti_path and os.path.exists(trasporti_path):
        print(f"Importing Transports from: {trasporti_path}")
        wb_trans = openpyxl.load_workbook(trasporti_path, read_only=True, data_only=True)
        transport_rows = []

        for sname in wb_trans.sheetnames:
            # Import sheets of year 2026
            if "2026" not in sname.upper():
                continue
            
            ws = wb_trans[sname]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or not row:
                    continue
                client_name = safe_str(row[3]) if len(row) > 3 else ""
                if not client_name or client_name.lower() == "cliente (destinazione)":
                    continue

                day_name = safe_str(row[0]) if len(row) > 0 else ""
                transport_date = safe_date_iso(row[1]) if len(row) > 1 else ""
                time_slot = safe_str(row[2]) if len(row) > 2 else ""
                city = safe_str(row[5]) if len(row) > 5 else ""
                province = safe_str(row[6]).upper() if len(row) > 6 else ""
                weight_kg = safe_float(row[7], 0.0) if len(row) > 7 else 0.0
                carrier = safe_str(row[8]) if len(row) > 8 else ""
                notes = safe_str(row[9]) if len(row) > 9 else ""
                zone = safe_str(row[10]) if len(row) > 10 else ""
                charge = safe_float(row[11], 0.0) if len(row) > 11 else 0.0

                transport_rows.append((
                    transport_date, day_name, time_slot, client_name,
                    city, province, weight_kg, carrier, notes, zone, charge, sname
                ))

        cursor.execute("DELETE FROM transports")
        cursor.executemany("""
            INSERT INTO transports (
                transport_date, day_name, time_slot, client_name,
                city, province, weight_kg, carrier, notes, zone, charge, sheet_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, transport_rows)
        total_transports = len(transport_rows)
        wb_trans.close()
        print(f"Imported {total_transports} transports.")

    # Record sync log
    cursor.execute("""
        INSERT INTO sync_logs (source, status, total_clients, total_articles, total_orders, total_prices, details)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, ("Excel Import", "SUCCESS", total_clients, total_articles, total_orders, len(price_map), f"Clienti: {total_clients}, Trasporti: {total_transports}"))

    conn.commit()
    conn.close()

    return {
        "status": "success",
        "clients": total_clients,
        "articles": total_articles,
        "orders": total_orders,
        "prices": len(price_map),
        "transports": total_transports,
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

if __name__ == "__main__":
    res = import_all_excel_data()
    print("Import Result:", res)
